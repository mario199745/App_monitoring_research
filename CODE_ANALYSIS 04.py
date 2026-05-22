# -*- coding: utf-8 -*-
"""
Edita el Excel de DATA para agregar o actualizar una columna consolidada
con solo dos categorias:

- Tesis
- Artículo científico

La columna se deriva desde "General_ Tipo de Publicación" y se aplica
unicamente sobre la hoja "REPOSITORIO_DEPURADO".
"""

from __future__ import annotations

import argparse
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL = BASE_DIR / "data" / "BD_08_04_2026_DEPURADO_20260408_015909.xlsx"
TARGET_SHEET = "REPOSITORIO_DEPURADO"
SOURCE_COLUMN = "General_ Tipo de Publicación"
TARGET_COLUMN = "Categoria_Tesis_Articulo"


def normalize_text(value) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text


def classify_publication_type(value) -> str:
    text = normalize_text(value)
    if "tesis" in text:
        return "Tesis"
    return "Artículo científico"


def create_backup(excel_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = excel_path.with_name(f"{excel_path.stem}_backup_{timestamp}{excel_path.suffix}")
    shutil.copy2(excel_path, backup_path)
    return backup_path


def update_excel(excel_path: Path, create_file_backup: bool = True) -> dict[str, int]:
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {excel_path}")

    if excel_path.name.startswith("~$"):
        raise ValueError("No se debe procesar un archivo temporal de Excel que empieza con '~$'.")

    backup_path = create_backup(excel_path) if create_file_backup else None

    try:
        workbook = load_workbook(excel_path)
        if TARGET_SHEET not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja requerida: {TARGET_SHEET}")

        worksheet = workbook[TARGET_SHEET]
        headers = [cell.value for cell in worksheet[1]]

        if SOURCE_COLUMN not in headers:
            raise ValueError(f"No existe la columna requerida: {SOURCE_COLUMN}")

        source_idx = headers.index(SOURCE_COLUMN) + 1
        target_idx = (
            headers.index(TARGET_COLUMN) + 1
            if TARGET_COLUMN in headers
            else worksheet.max_column + 1
        )

        worksheet.cell(row=1, column=target_idx, value=TARGET_COLUMN)

        counts = {"Tesis": 0, "Artículo científico": 0}
        for row_idx in range(2, worksheet.max_row + 1):
            source_value = worksheet.cell(row=row_idx, column=source_idx).value
            category = classify_publication_type(source_value)
            worksheet.cell(row=row_idx, column=target_idx, value=category)
            counts[category] += 1

        workbook.save(excel_path)
        workbook.close()

    except PermissionError as exc:
        message = (
            "No se pudo guardar el Excel. Cierra el archivo en Excel u otra aplicacion "
            "y vuelve a ejecutar el script."
        )
        raise PermissionError(message) from exc

    if backup_path is not None:
        print(f"Backup creado: {backup_path}")

    return counts


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Agrega o actualiza la columna Categoria_Tesis_Articulo en el Excel de DATA."
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL,
        help=f"Ruta del Excel a editar. Por defecto: {DEFAULT_EXCEL}",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="No crear copia de seguridad antes de modificar el Excel.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    counts = update_excel(args.excel, create_file_backup=not args.no_backup)

    print("Excel actualizado correctamente.")
    print(f"Hoja procesada: {TARGET_SHEET}")
    print(f"Columna origen: {SOURCE_COLUMN}")
    print(f"Columna generada: {TARGET_COLUMN}")
    print("Conteo final:")
    for category, count in counts.items():
        print(f"- {category}: {count}")


if __name__ == "__main__":
    main()
