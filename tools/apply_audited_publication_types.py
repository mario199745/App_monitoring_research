from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ID = "ID_PUBLICACION_PROPUESTA"
TYPE = "TIPO_PUBLICACION_PUBLICO"
SUBTYPE = "SUBTIPO_PUBLICACION_PUBLICO"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Crea una copia del maestro y aplica la taxonomía auditada."
    )
    parser.add_argument("--master", type=Path, required=True)
    parser.add_argument("--audit", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--decisions", type=Path, required=True)
    parser.add_argument(
        "--group-events-as-articles",
        action="store_true",
        help=(
            "Agrupa las publicaciones de evento en Artículo y usa "
            "Publicación de evento científico como subtipo."
        ),
    )
    args = parser.parse_args()

    audit = pd.read_excel(
        args.audit, sheet_name="AUDITORIA_COMPLETA", dtype=object, engine="openpyxl"
    )
    required = {ID, "TIPO_PROPUESTO", "SUBTIPO_PROPUESTO"}
    missing = required - set(audit.columns)
    if missing:
        raise ValueError(f"Faltan columnas en la auditoría: {sorted(missing)}")
    decisions = audit[[ID, "TIPO_PROPUESTO", "SUBTIPO_PROPUESTO"]].copy()
    if args.group_events_as_articles:
        event_mask = (
            decisions["TIPO_PROPUESTO"].eq("Publicación de evento científico")
            | decisions["SUBTIPO_PROPUESTO"].isin(
                ["Artículo de conferencia", "Ponencia o memoria de evento"]
            )
        )
        decisions.loc[event_mask, "TIPO_PROPUESTO"] = "Artículo"
        decisions.loc[event_mask, "SUBTIPO_PROPUESTO"] = (
            "Publicación de evento científico"
        )
    if decisions[ID].isna().any() or decisions[ID].duplicated().any():
        raise ValueError("Los IDs de la auditoría no son completos y únicos.")
    if decisions[["TIPO_PROPUESTO", "SUBTIPO_PROPUESTO"]].isna().any().any():
        raise ValueError("Existen decisiones sin tipo o subtipo.")

    args.decisions.parent.mkdir(parents=True, exist_ok=True)
    decisions.to_csv(args.decisions, index=False, encoding="utf-8-sig")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output_dir / f"BD_APP_FINAL_{stamp}_DEDUPLICADA.xlsx"
    shutil.copy2(args.master, output)
    workbook = load_workbook(output)
    sheet = workbook["BD_APP"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    missing_headers = {ID, TYPE, SUBTYPE} - set(headers)
    if missing_headers:
        raise ValueError(f"Faltan columnas en BD_APP: {sorted(missing_headers)}")

    decision_map = decisions.set_index(ID).to_dict(orient="index")
    seen: set[str] = set()
    for row in range(2, sheet.max_row + 1):
        publication_id = str(sheet.cell(row, headers[ID]).value)
        decision = decision_map.get(publication_id)
        if decision is None:
            raise ValueError(f"No existe decisión para {publication_id}")
        sheet.cell(row, headers[TYPE]).value = decision["TIPO_PROPUESTO"]
        sheet.cell(row, headers[SUBTYPE]).value = decision["SUBTIPO_PROPUESTO"]
        seen.add(publication_id)
    unused = set(decision_map) - seen
    if unused:
        raise ValueError(f"Existen decisiones no aplicadas: {len(unused)}")

    trace_name = "TRAZABILIDAD_TIPOS"
    if trace_name in workbook.sheetnames:
        del workbook[trace_name]
    trace = workbook.create_sheet(trace_name)
    trace.append(["INDICADOR", "VALOR"])
    trace.append(["fecha_aplicacion", datetime.now().isoformat(timespec="seconds")])
    trace.append(["maestro_origen", args.master.name])
    trace.append(["auditoria_origen", args.audit.name])
    trace.append(["publicaciones_actualizadas", len(decisions)])
    trace.append(["columnas_actualizadas", f"{TYPE}; {SUBTYPE}"])
    for publication_type, count in decisions["TIPO_PROPUESTO"].value_counts().items():
        trace.append([f"tipo_{publication_type}", int(count)])

    workbook.save(output)
    print(f"OUTPUT={output.resolve()}")
    print(f"DECISIONS={args.decisions.resolve()}")
    print(decisions["TIPO_PROPUESTO"].value_counts().to_string())
    print(decisions["SUBTIPO_PROPUESTO"].value_counts().to_string())


if __name__ == "__main__":
    main()
