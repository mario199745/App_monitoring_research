from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ID = "ID_PUBLICACION_PROPUESTA"
TYPE = "TIPO_PUBLICACION_PUBLICO"
SUBTYPE = "SUBTIPO_PUBLICACION_PUBLICO"
LEVEL = "NIVEL_ACADEMICO_PUBLICO"
SUMMARY = "General_ Resumen"
DETAIL = "DETALLE_TESIS_POSGRADO_PUBLICO"

MASTER_PATTERN = re.compile(
    r"\b(tesis de maestr[ií]a|master[’']*s? thesis|master[’']*s? dissertation|"
    r"msc thesis|m\.sc\. thesis|grado de mag[ií]ster|degree of master|"
    r"maestr[ií]a|mag[ií]ster)\b",
    re.IGNORECASE,
)
DOCTORAL_PATTERN = re.compile(
    r"\b(tesis doctoral|doctoral thesis|doctoral dissertation|"
    r"ph\.?\s*d\.? (?:thesis|dissertation)|doctor of philosophy|"
    r"grado de doctor|degree of doctor|doctorado)\b",
    re.IGNORECASE,
)


def classify_detail(row: pd.Series) -> object:
    if str(row.get(TYPE, "")).strip() != "Tesis":
        return pd.NA
    subtype = str(row.get(SUBTYPE, "")).strip()
    level = str(row.get(LEVEL, "")).strip()
    summary = "" if pd.isna(row.get(SUMMARY)) else str(row.get(SUMMARY))
    if subtype == "Tesis de maestría" or level == "Maestría":
        return "Tesis de maestría"
    if subtype == "Tesis doctoral" or level == "Doctorado":
        return "Tesis doctoral"
    if MASTER_PATTERN.search(summary):
        return "Tesis de maestría"
    if DOCTORAL_PATTERN.search(summary):
        return "Tesis doctoral"
    if subtype == "Tesis de posgrado":
        return "No identificados"
    return pd.NA


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--master", type=Path, required=True)
    parser.add_argument("--decisions", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()

    data = pd.read_excel(args.master, sheet_name="BD_APP", dtype=object)
    data[DETAIL] = data.apply(classify_detail, axis=1)
    postgraduate = data[DETAIL].notna()
    data.loc[postgraduate, SUBTYPE] = "Tesis de posgrado"
    data.loc[data[DETAIL].eq("Tesis de maestría"), LEVEL] = "Maestría"
    data.loc[data[DETAIL].eq("Tesis doctoral"), LEVEL] = "Doctorado"
    data.loc[data[DETAIL].eq("No identificados"), LEVEL] = "No identificado"

    counts = data.loc[postgraduate, DETAIL].value_counts()
    expected = {
        "Tesis de maestría": 54,
        "Tesis doctoral": 24,
        "No identificados": 389,
    }
    if counts.to_dict() != expected:
        raise ValueError(f"Distribución inesperada: {counts.to_dict()}")

    decisions = pd.read_csv(args.decisions, dtype="string", encoding="utf-8-sig")
    updates = data[[ID, TYPE, SUBTYPE, DETAIL]].rename(
        columns={TYPE: "TIPO_PROPUESTO", SUBTYPE: "SUBTIPO_PROPUESTO"}
    )
    decisions = decisions.drop(
        columns=["TIPO_PROPUESTO", "SUBTIPO_PROPUESTO", DETAIL],
        errors="ignore",
    ).merge(updates, on=ID, how="left", validate="one_to_one")
    decisions.to_csv(args.decisions, index=False, encoding="utf-8-sig")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output_dir / f"BD_APP_FINAL_{stamp}_DEDUPLICADA.xlsx"
    shutil.copy2(args.master, output)
    workbook = load_workbook(output)
    sheet = workbook["BD_APP"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    if DETAIL not in headers:
        detail_column = sheet.max_column + 1
        sheet.cell(1, detail_column).value = DETAIL
        headers[DETAIL] = detail_column
    update_map = data.set_index(ID)[[SUBTYPE, LEVEL, DETAIL]].to_dict(orient="index")
    for row_number in range(2, sheet.max_row + 1):
        publication_id = str(sheet.cell(row_number, headers[ID]).value)
        values = update_map[publication_id]
        for column in [SUBTYPE, LEVEL, DETAIL]:
            value = values[column]
            sheet.cell(row_number, headers[column]).value = (
                None if pd.isna(value) else value
            )

    trace = workbook["TRAZABILIDAD_TIPOS"]
    trace.append(["jerarquia_posgrado_aplicada", datetime.now().isoformat(timespec="seconds")])
    for label, count in counts.items():
        trace.append([f"detalle_posgrado_{label}", int(count)])
    workbook.save(output)

    print(f"OUTPUT={output.resolve()}")
    print(counts.to_string())


if __name__ == "__main__":
    main()
